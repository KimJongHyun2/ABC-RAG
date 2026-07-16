import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, PieChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

df = pd.read_csv("data/yes24_it_mobile_bestsellers.csv")
df.columns = df.columns.str.strip()
df["Rating"] = pd.to_numeric(df["Rating"], errors="coerce")
df["ReviewCount"] = pd.to_numeric(df["ReviewCount"], errors="coerce")
df["Price"] = pd.to_numeric(df["Price"], errors="coerce")

wb = Workbook()

# ── Colors & Styles ──
DARK_BLUE = "1F4E79"
MID_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
ACCENT_GREEN = "548235"
ACCENT_ORANGE = "ED7D31"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"

header_font = Font(name="Arial", bold=True, color=WHITE, size=11)
header_fill = PatternFill("solid", fgColor=DARK_BLUE)
sub_header_font = Font(name="Arial", bold=True, color=WHITE, size=10)
sub_header_fill = PatternFill("solid", fgColor=MID_BLUE)
title_font = Font(name="Arial", bold=True, size=14, color=DARK_BLUE)
metric_font = Font(name="Arial", bold=True, size=20, color=DARK_BLUE)
metric_label = Font(name="Arial", size=9, color="666666")
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
center = Alignment(horizontal="center", vertical="center")
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header_row(ws, row, cols, font=header_font, fill=header_fill):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = center
        cell.border = thin_border

def style_data_cell(cell, is_alt=False):
    cell.font = Font(name="Arial", size=10)
    cell.border = thin_border
    if is_alt:
        cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)

# ════════════════════════════════════════════
# Sheet 1: 대시보드 요약
# ════════════════════════════════════════════
ws = wb.active
ws.title = "대시보드"
ws.sheet_properties.tabColor = DARK_BLUE

ws.merge_cells("A1:H1")
ws["A1"] = "YES24 IT/모바일 베스트셀러 대시보드"
ws["A1"].font = Font(name="Arial", bold=True, size=18, color=DARK_BLUE)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 40

ws.merge_cells("A2:H2")
ws["A2"] = f"총 {len(df)}권 분석 | 데이터 기준: YES24 IT 모바일 종합 베스트셀러"
ws["A2"].font = Font(name="Arial", size=10, color="888888")
ws["A2"].alignment = Alignment(horizontal="center")

# ── 핵심 지표 ──
metrics = [
    ("총 도서 수", f"{len(df)}권", "A4"),
    ("평균 가격", f"{df['Price'].mean():,.0f}원", "C4"),
    ("평균 평점", f"{df['Rating'].mean():.1f}", "E4"),
    ("총 리뷰 수", f"{df['ReviewCount'].sum():,.0f}", "G4"),
]
for label, value, cell_ref in metrics:
    r = ws[cell_ref]
    r.value = value
    r.font = metric_font
    r.alignment = center
    r.border = thin_border
    r.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    lr = ws[cell_ref.replace("4", "5")]
    lr.value = label
    lr.font = metric_label
    lr.alignment = center

for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
    ws.column_dimensions[col].width = 16

# ── 출판사별 도서 수 Top 10 ──
ws.merge_cells("A7:D7")
ws["A7"] = "출판사별 도서 수 Top 10"
ws["A7"].font = Font(name="Arial", bold=True, size=12, color=DARK_BLUE)

pub_counts = df["Publisher"].value_counts().head(10).reset_index()
pub_counts.columns = ["Publisher", "Count"]

ws.cell(row=8, column=1, value="출판사")
ws.cell(row=8, column=2, value="도서 수")
ws.cell(row=8, column=3, value="비율")
style_header_row(ws, 8, 3, sub_header_font, sub_header_fill)

for i, (_, row) in enumerate(pub_counts.iterrows()):
    r = 9 + i
    ws.cell(row=r, column=1, value=row["Publisher"])
    ws.cell(row=r, column=2, value=row["Count"])
    c3 = ws.cell(row=r, column=3)
    c3.value = f'=B{r}/SUM(B9:B18)'
    c3.number_format = '0.0%'
    for c in range(1, 4):
        style_data_cell(ws.cell(row=r, column=c), i % 2 == 0)

# ── 가격대별 분포 ──
ws.merge_cells("F7:H7")
ws["F7"] = "가격대별 도서 수"
ws["F7"].font = Font(name="Arial", bold=True, size=12, color=DARK_BLUE)

bins = [0, 15000, 20000, 25000, 30000, 40000, 100000]
labels = ["~1.5만", "1.5~2만", "2~2.5만", "2.5~3만", "3~4만", "4만+"]
df["PriceBin"] = pd.cut(df["Price"], bins=bins, labels=labels)
price_dist = df["PriceBin"].value_counts().reindex(labels).fillna(0).astype(int)

ws.cell(row=8, column=6, value="가격대")
ws.cell(row=8, column=7, value="도서 수")
style_header_row(ws, 8, 7, sub_header_font, sub_header_fill)
ws.cell(row=8, column=8).font = sub_header_font
ws.cell(row=8, column=8).fill = sub_header_fill

for i, label in enumerate(labels):
    r = 9 + i
    ws.cell(row=r, column=6, value=label)
    ws.cell(row=r, column=7, value=int(price_dist[label]))
    for c in [6, 7]:
        style_data_cell(ws.cell(row=r, column=c), i % 2 == 0)

# ── 리뷰 수 Top 10 ──
ws.merge_cells("A21:D21")
ws["A21"] = "리뷰 수 Top 10 도서"
ws["A21"].font = Font(name="Arial", bold=True, size=12, color=DARK_BLUE)

top_review = df.nlargest(10, "ReviewCount")
headers = ["순위", "제목", "저자", "리뷰 수"]
for c, h in enumerate(headers, 1):
    ws.cell(row=22, column=c, value=h)
style_header_row(ws, 22, 4, sub_header_font, sub_header_fill)

for i, (_, row) in enumerate(top_review.iterrows()):
    r = 23 + i
    ws.cell(row=r, column=1, value=int(row["Rank"]))
    ws.cell(row=r, column=2, value=row["Title"])
    ws.cell(row=r, column=3, value=row["Author"])
    ws.cell(row=r, column=4, value=int(row["ReviewCount"]))
    for c in range(1, 5):
        style_data_cell(ws.cell(row=r, column=c), i % 2 == 0)
    ws.cell(row=r, column=2).alignment = left_wrap

ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 45
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 12

# ════════════════════════════════════════════
# Sheet 2: 차트
# ════════════════════════════════════════════
ws2 = wb.create_sheet("차트")
ws2.sheet_properties.tabColor = MID_BLUE

# ── 출판사별 도서 수 차트용 데이터 ──
ws2["A1"] = "출판사"
ws2["B1"] = "도서 수"
for i, (_, row) in enumerate(pub_counts.iterrows()):
    ws2.cell(row=i + 2, column=1, value=row["Publisher"])
    ws2.cell(row=i + 2, column=2, value=row["Count"])

chart1 = BarChart()
chart1.type = "bar"
chart1.title = "출판사별 도서 수 Top 10"
chart1.y_axis.title = "출판사"
chart1.x_axis.title = "도서 수"
chart1.style = 10
chart1.width = 28
chart1.height = 16
data1 = Reference(ws2, min_col=2, min_row=1, max_row=11)
cats1 = Reference(ws2, min_col=1, min_row=2, max_row=11)
chart1.add_data(data1, titles_from_data=True)
chart1.set_categories(cats1)
chart1.shape = 4
s = chart1.series[0]
s.graphicalProperties.solidFill = MID_BLUE
ws2.add_chart(chart1, "D1")

# ── 가격대별 분포 차트용 데이터 ──
ws2["A16"] = "가격대"
ws2["B16"] = "도서 수"
for i, label in enumerate(labels):
    ws2.cell(row=17 + i, column=1, value=label)
    ws2.cell(row=17 + i, column=2, value=int(price_dist[label]))

chart2 = BarChart()
chart2.type = "col"
chart2.title = "가격대별 도서 수 분포"
chart2.y_axis.title = "도서 수"
chart2.x_axis.title = "가격대"
chart2.style = 10
chart2.width = 22
chart2.height = 14
data2 = Reference(ws2, min_col=2, min_row=16, max_row=22)
cats2 = Reference(ws2, min_col=1, min_row=17, max_row=22)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
s2 = chart2.series[0]
s2.graphicalProperties.solidFill = ACCENT_ORANGE
ws2.add_chart(chart2, "D20")

# ── 평점 분포 ──
rating_bins = [0, 8, 9, 9.5, 9.8, 10.1]
rating_labels = ["8.0~8.9", "9.0~9.4", "9.5~9.7", "9.8~9.9", "10.0"]
df["RatingBin"] = pd.cut(df["Rating"], bins=rating_bins, labels=rating_labels, right=True)
rating_dist = df["RatingBin"].value_counts().reindex(rating_labels).fillna(0).astype(int)
rating_dist_with_na = pd.concat([rating_dist, pd.Series({"결측": df["Rating"].isna().sum()})])

ws2["A25"] = "평점 구간"
ws2["B25"] = "도서 수"
for i, (label, count) in enumerate(rating_dist_with_na.items()):
    ws2.cell(row=26 + i, column=1, value=label)
    ws2.cell(row=26 + i, column=2, value=int(count))

chart3 = PieChart()
chart3.title = "평점 구간별 분포"
chart3.style = 10
chart3.width = 18
chart3.height = 14
data3 = Reference(ws2, min_col=2, min_row=25, max_row=31)
cats3 = Reference(ws2, min_col=1, min_row=26, max_row=31)
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
chart3.dataLabels = DataLabelList()
chart3.dataLabels.showPercent = True
chart3.dataLabels.showVal = True
ws2.add_chart(chart3, "D38")

# ── 가격 vs 평점 (BAR 차트로 대체) ──
ws2["A34"] = "평점 구간"
ws2["B34"] = "평균 가격"
rating_price = df.dropna(subset=["Rating"]).copy()
rating_price["RBin"] = pd.cut(rating_price["Rating"], bins=[0, 9, 9.5, 9.8, 10.1], labels=["~9.0", "9.0~9.5", "9.5~9.8", "9.8~10.0"])
rp_avg = rating_price.groupby("RBin", observed=True)["Price"].mean()

for i, (label, avg) in enumerate(rp_avg.items()):
    ws2.cell(row=35 + i, column=1, value=label)
    ws2.cell(row=35 + i, column=2, value=round(avg))

chart4 = BarChart()
chart4.type = "col"
chart4.title = "평점 구간별 평균 가격"
chart4.y_axis.title = "평균 가격 (원)"
chart4.style = 10
chart4.width = 22
chart4.height = 14
data4 = Reference(ws2, min_col=2, min_row=34, max_row=38)
cats4 = Reference(ws2, min_col=1, min_row=35, max_row=38)
chart4.add_data(data4, titles_from_data=True)
chart4.set_categories(cats4)
s4 = chart4.series[0]
s4.graphicalProperties.solidFill = ACCENT_GREEN
ws2.add_chart(chart4, "J20")

for col in ["A", "B", "C"]:
    ws2.column_dimensions[col].width = 18

# ════════════════════════════════════════════
# Sheet 3: 전체 데이터
# ════════════════════════════════════════════
ws3 = wb.create_sheet("전체 데이터")
ws3.sheet_properties.tabColor = ACCENT_GREEN

cols = ["Rank", "Title", "Author", "Publisher", "Price", "Rating", "ReviewCount"]
col_widths = [8, 50, 25, 20, 12, 10, 12]

for c, (col_name, w) in enumerate(zip(cols, col_widths), 1):
    ws3.cell(row=1, column=c, value=col_name)
    ws3.column_dimensions[get_column_letter(c)].width = w
style_header_row(ws3, 1, len(cols))

for i, (_, row) in enumerate(df[cols].iterrows()):
    r = i + 2
    for c, col_name in enumerate(cols, 1):
        val = row[col_name]
        cell = ws3.cell(row=r, column=c)
        if col_name in ["Rank", "ReviewCount"] and pd.notna(val):
            cell.value = int(val)
        elif col_name == "Price" and pd.notna(val):
            cell.value = int(val)
            cell.number_format = '#,##0'
        elif col_name == "Rating" and pd.notna(val):
            cell.value = round(val, 1)
        else:
            cell.value = val
        style_data_cell(cell, i % 2 == 0)
        if col_name == "Title":
            cell.alignment = left_wrap

ws3.auto_filter.ref = f"A1:G{len(df) + 1}"
ws3.freeze_panes = "A2"

# ════════════════════════════════════════════
# Sheet 4: 출판사별 상세
# ════════════════════════════════════════════
ws4 = wb.create_sheet("출판사별 분석")
ws4.sheet_properties.tabColor = ACCENT_ORANGE

ws4.merge_cells("A1:F1")
ws4["A1"] = "출판사별 상세 분석"
ws4["A1"].font = title_font
ws4["A1"].alignment = Alignment(horizontal="center")
ws4.row_dimensions[1].height = 30

pub_stats = (
    df.groupby("Publisher")
    .agg(
        도서수=("Rank", "count"),
        평균평점=("Rating", "mean"),
        평균가격=("Price", "mean"),
        총리뷰=("ReviewCount", "sum"),
        최고평점=("Rating", "max"),
    )
    .query("도서수 >= 3")
    .sort_values("도서수", ascending=False)
    .reset_index()
)

headers4 = ["출판사", "도서 수", "평균 평점", "평균 가격", "총 리뷰", "최고 평점"]
for c, h in enumerate(headers4, 1):
    ws4.cell(row=3, column=c, value=h)
style_header_row(ws4, 3, 6, sub_header_font, sub_header_fill)

for i, (_, row) in enumerate(pub_stats.iterrows()):
    r = 4 + i
    ws4.cell(row=r, column=1, value=row["Publisher"])
    ws4.cell(row=r, column=2, value=int(row["도서수"]))
    c3 = ws4.cell(row=r, column=3)
    c3.value = round(row["평균평점"], 1) if pd.notna(row["평균평점"]) else "-"
    c4 = ws4.cell(row=r, column=4)
    c4.value = int(row["평균가격"]) if pd.notna(row["평균가격"]) else "-"
    c4.number_format = '#,##0'
    ws4.cell(row=r, column=5, value=int(row["총리뷰"]) if pd.notna(row["총리뷰"]) else 0)
    ws4.cell(row=r, column=6, value=row["최고평점"] if pd.notna(row["최고평점"]) else "-")
    for c in range(1, 7):
        style_data_cell(ws4.cell(row=r, column=c), i % 2 == 0)

ws4.column_dimensions["A"].width = 22
ws4.column_dimensions["B"].width = 10
ws4.column_dimensions["C"].width = 12
ws4.column_dimensions["D"].width = 14
ws4.column_dimensions["E"].width = 12
ws4.column_dimensions["F"].width = 12

# ── 저장 ──
output_path = "data/yes24_it_mobile_bestsellers_dashboard.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
